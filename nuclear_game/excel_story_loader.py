# -*- coding: utf-8 -*-
"""从 Excel 读取剧本表，并提供面向剧本人员的详细格式检查。"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


EFFECT_KEY_ALIASES = {
    "冷凝器绝对压力": "condenser_pressure",
    "冷凝器压力": "condenser_pressure",
    "一回路流量": "primary_flow",
    "主泵流量": "primary_flow",
    "蒸汽发生器水位": "sg_level",
    "一回路平均温度": "temperature",
    "平均温度": "temperature",
    "资金": "funds",
    "安全评分": "safety",
    "防护评分": "protection_score",
    "冷却水流量": "cooling_flow",
}
KNOWN_EFFECT_KEYS = set(EFFECT_KEY_ALIASES.values())
REQUIRED_STAGE_HEADERS = ["阶段", "章节", "标题", "主线任务", "说明", "目标", "任务", "奖励", "知识复盘", "通关语"]
REQUIRED_STAGE_FIELDS = ["阶段", "标题", "主线任务", "说明", "目标", "任务", "通关语"]
LEVEL_SHEET_NAMES = ("关卡表", "关卡填写表")
OPTIONAL_SHEETS = ("任务表", "知识点表", "事故卡表", "奖励表")


def _candidate_paths() -> List[Path]:
    names = ["剧本填写表.xlsx", "story_table.xlsx"]
    roots = []
    roots.append(Path(__file__).resolve().parent.parent)
    roots.append(Path.cwd())
    if getattr(sys, "frozen", False):
        roots.append(Path(sys.executable).resolve().parent)
    seen = []
    for root in roots:
        for name in names:
            path = root / name
            if path not in seen:
                seen.append(path)
    return seen


def _split_lines(value: Any) -> List[str]:
    if value is None:
        return []
    text = str(value).replace("\r\n", "\n").replace("；", "\n")
    return [line.strip() for line in text.split("\n") if line.strip()]


def _normalize_effect_key(key: str) -> str:
    key = key.strip()
    return EFFECT_KEY_ALIASES.get(key, key)


def _parse_number(value: str) -> Optional[float]:
    try:
        return float(str(value).strip().replace("+", ""))
    except Exception:
        return None


def _parse_effect(value: Any) -> Dict[str, str]:
    effect: Dict[str, str] = {}
    for line in _split_lines(value):
        if "=" in line:
            k, v = line.split("=", 1)
        elif "：" in line:
            k, v = line.split("：", 1)
        elif ":" in line:
            k, v = line.split(":", 1)
        else:
            continue
        key = _normalize_effect_key(k)
        # 只把可识别且可转成数字的影响写入游戏，避免剧本错误导致运行崩溃。
        if key in KNOWN_EFFECT_KEYS and _parse_number(v) is not None:
            effect[key] = v.strip()
    return effect


def _parse_objectives(value: Any) -> List[dict]:
    result = []
    for line in _split_lines(value):
        if "|" in line:
            name, hint = line.split("|", 1)
        elif "：" in line:
            name, hint = line.split("：", 1)
        else:
            name, hint = line, ""
        if name.strip():
            result.append({"名称": name.strip(), "提示": hint.strip()})
    return result


def _open_workbook(path: Path):
    from openpyxl import load_workbook  # type: ignore
    return load_workbook(path, data_only=True)


def _row_dicts(ws) -> List[Tuple[int, Dict[str, Any]]]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    result: List[Tuple[int, Dict[str, Any]]] = []
    for excel_row, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
        if any(value is not None and str(value).strip() for value in item.values()):
            result.append((excel_row, item))
    return result


def _stage_value(item: Dict[str, Any]) -> Optional[int]:
    try:
        return int(item.get("阶段"))
    except Exception:
        return None


def _merge_optional_sheets(wb, rows: List[dict]) -> dict:
    """把任务表、知识点表、奖励表等多 Sheet 内容合并到关卡行。"""
    by_stage = {row["阶段"]: row for row in rows}
    extra = {"事故卡表": []}

    if "任务表" in wb.sheetnames:
        for _, item in _row_dicts(wb["任务表"]):
            stage = _stage_value(item)
            if stage in by_stage:
                name = str(item.get("目标") or item.get("名称") or item.get("任务") or "").strip()
                hint = str(item.get("提示") or item.get("说明") or "").strip()
                task_text = str(item.get("任务") or name).strip()
                if task_text and task_text not in by_stage[stage]["任务"]:
                    by_stage[stage]["任务"].append(task_text)
                if name:
                    by_stage[stage]["目标"].append({"名称": name, "提示": hint})

    if "知识点表" in wb.sheetnames:
        for _, item in _row_dicts(wb["知识点表"]):
            stage = _stage_value(item)
            text = str(item.get("知识点") or item.get("知识复盘") or item.get("内容") or "").strip()
            if stage in by_stage and text and text not in by_stage[stage]["知识复盘"]:
                by_stage[stage]["知识复盘"].append(text)

    if "奖励表" in wb.sheetnames:
        for _, item in _row_dicts(wb["奖励表"]):
            stage = _stage_value(item)
            if stage not in by_stage:
                continue
            for field in ("奖励", "解锁"):
                for value in _split_lines(item.get(field)):
                    if value not in by_stage[stage][field]:
                        by_stage[stage][field].append(value)
            badge = str(item.get("资质章") or "").strip()
            if badge:
                by_stage[stage]["资质章"] = badge

    if "事故卡表" in wb.sheetnames:
        for _, item in _row_dicts(wb["事故卡表"]):
            if item.get("事故名称") or item.get("现象"):
                extra["事故卡表"].append({str(k): str(v) for k, v in item.items() if v is not None})

    return extra


def _load_xlsx(path: Path) -> Optional[dict]:
    try:
        from openpyxl import load_workbook  # noqa: F401
    except Exception:
        return None
    if not path.exists():
        return None
    ok, _messages = check_excel_story(path)
    if not ok:
        return None
    wb = _open_workbook(path)

    title = None
    subtitle = None
    objective = None
    if "游戏设置" in wb.sheetnames:
        ws = wb["游戏设置"]
        data = {str(row[0]).strip(): row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
        title = data.get("游戏标题")
        subtitle = data.get("游戏副标题")
        objective = {
            "玩家身份": data.get("玩家身份", "核电项目负责人"),
            "总目标": data.get("总目标", "完成五阶段建设并安全并网。"),
            "胜利条件": _split_lines(data.get("胜利条件", "")),
            "失败条件": _split_lines(data.get("失败条件", "")),
        }

    level_sheet_name = next((name for name in LEVEL_SHEET_NAMES if name in wb.sheetnames), None)
    if not level_sheet_name:
        return None
    ws = wb[level_sheet_name]
    rows = []
    for _, item in _row_dicts(ws):
        if item.get("阶段") is None or item.get("标题") is None:
            continue
        stage = _stage_value(item)
        if stage is None:
            continue
        rows.append({
            "阶段": stage,
            "章节": str(item.get("章节") or ""),
            "标题": str(item.get("标题") or ""),
            "主线任务": str(item.get("主线任务") or ""),
            "说明": _split_lines(item.get("说明")),
            "特殊规则": str(item.get("特殊规则") or ""),
            "影响": _parse_effect(item.get("影响")),
            "目标": _parse_objectives(item.get("目标")),
            "任务": _split_lines(item.get("任务")),
            "奖励": _split_lines(item.get("奖励")),
            "解锁": _split_lines(item.get("解锁")),
            "资质章": str(item.get("资质章") or ""),
            "知识复盘": _split_lines(item.get("知识复盘")),
            "通关语": str(item.get("通关语") or ""),
        })
    if not rows:
        return None
    extra = _merge_optional_sheets(wb, rows)
    return {"游戏标题": title, "游戏副标题": subtitle, "新手目标": objective,
            "关卡填写表": rows, "事故卡表": extra.get("事故卡表", []),
            "source": str(path), "level_sheet": level_sheet_name}


def check_excel_story(path: Optional[Path] = None) -> Tuple[bool, List[str]]:
    """详细检查 Excel 剧本表，返回 (是否通过, 消息列表)。

    这个函数专门给 tools/Excel剧本检查.py 使用，错误会定位到行号和字段名。
    """
    messages: List[str] = []
    paths = [path] if path else _candidate_paths()
    target = next((Path(p) for p in paths if p and Path(p).exists()), None)
    if target is None:
        messages.append("未找到剧本填写表.xlsx 或 story_table.xlsx，将使用 story_plain.py 备用剧本。")
        return True, messages
    try:
        wb = _open_workbook(target)
    except Exception as exc:
        return False, [f"无法打开 Excel：{target}", f"原因：{exc}"]

    messages.append(f"正在检查：{target}")
    if "游戏设置" not in wb.sheetnames:
        messages.append("提示：未找到【游戏设置】工作表，将使用默认游戏标题和新手目标。")
    level_sheet_name = next((name for name in LEVEL_SHEET_NAMES if name in wb.sheetnames), None)
    if level_sheet_name is None:
        return False, messages + ["错误：缺少【关卡表】或【关卡填写表】工作表。"]

    ws = wb[level_sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    header_set = set(headers)
    missing_headers = [h for h in REQUIRED_STAGE_HEADERS if h not in header_set]
    if missing_headers:
        messages.append("错误：关卡填写表缺少表头：" + "、".join(missing_headers))

    stages = []
    row_count = 0
    for excel_row, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
        if not any(value is not None and str(value).strip() for value in item.values()):
            continue
        row_count += 1
        for field in REQUIRED_STAGE_FIELDS:
            if field in header_set and (item.get(field) is None or not str(item.get(field)).strip()):
                messages.append(f"错误：第 {excel_row} 行【{field}】不能为空。")
        try:
            stage = int(item.get("阶段"))
            stages.append(stage)
            if stage < 0 or stage > 4:
                messages.append(f"错误：第 {excel_row} 行【阶段】应为 0-4，目前是 {stage}。")
        except Exception:
            messages.append(f"错误：第 {excel_row} 行【阶段】必须是数字 0-4。")

        for line in _split_lines(item.get("影响")):
            if not any(sep in line for sep in ("=", "：", ":")):
                messages.append(f"警告：第 {excel_row} 行【影响】格式可能错误：{line}。建议写成 冷凝器绝对压力=+1.5")
                continue
            sep = "=" if "=" in line else "：" if "：" in line else ":"
            key, val = line.split(sep, 1)
            key = _normalize_effect_key(key)
            numeric_value = _parse_number(val)
            if key in KNOWN_EFFECT_KEYS and numeric_value is None:
                messages.append(f"错误：第 {excel_row} 行【影响】的数值不是数字：{line}。")
            elif key not in KNOWN_EFFECT_KEYS and numeric_value is not None:
                messages.append(f"警告：第 {excel_row} 行【影响】中的参数名不识别：{key}，该数值影响不会写入游戏参数。")
            # 非数字的“影响”内容视为剧情说明，不报错，便于剧本人员写文字说明。

    if row_count == 0:
        messages.append(f"错误：【{level_sheet_name}】没有有效关卡行。")
    duplicates = sorted({x for x in stages if stages.count(x) > 1})
    if duplicates:
        messages.append("错误：阶段编号重复：" + "、".join(map(str, duplicates)))
    expected = set(range(5))
    got = set(stages)
    if got != expected:
        messages.append(f"错误：阶段编号必须覆盖 0-4；当前为 {sorted(got)}。")

    # 多 Sheet 维护提示与轻量校验。
    for sheet_name in OPTIONAL_SHEETS:
        if sheet_name not in wb.sheetnames:
            messages.append(f"提示：未找到【{sheet_name}】工作表，可继续运行；需要团队分工时建议补充。")
    if "任务表" in wb.sheetnames:
        headers2 = [str(cell.value).strip() if cell.value is not None else "" for cell in wb["任务表"][1]]
        for required in ("阶段", "任务"):
            if required not in headers2:
                messages.append(f"错误：【任务表】缺少表头【{required}】。")
    if "知识点表" in wb.sheetnames:
        headers2 = [str(cell.value).strip() if cell.value is not None else "" for cell in wb["知识点表"][1]]
        if "阶段" not in headers2 or not any(h in headers2 for h in ("知识点", "知识复盘", "内容")):
            messages.append("错误：【知识点表】至少需要【阶段】和【知识点/知识复盘/内容】表头。")

    has_error = any(msg.startswith("错误") for msg in messages)
    if not has_error:
        messages.append(f"检查通过：共读取 {row_count} 个关卡，阶段覆盖 0-4。")
    return not has_error, messages


def load_excel_story() -> Optional[dict]:
    for path in _candidate_paths():
        data = _load_xlsx(path)
        if data:
            return data
    return None
